from ast import match_case
import os
import pandas as pd
import zipfile
from tkinter import Tk, filedialog, Label, Button, StringVar, Frame, messagebox
from tkinter.ttk import Style
from datetime import datetime
import tempfile
from openpyxl import load_workbook

# Función para formatear fechas
def parse_date(date_str):
    date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y%m%d']
    for fmt in date_formats:
        try:
            return datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None

# Función para procesar el archivo Excel
def procesar_excel(ruta_archivo):
    try:
        df = pd.read_excel(ruta_archivo)
        columnas_necesarias = ['DATE', 'TIME', 'SESSION GUID']
        if not all(col in df.columns for col in columnas_necesarias):
            messagebox.showerror("Error", "El archivo no contiene las columnas necesarias")
            return None

        df['DATE'] = df['DATE'].astype(str).str.replace("/", "-")
        df['TIME'] = df['TIME'].astype(str).str.replace(":", "")
        df['MATCH NAME'] = 'chat_' + df['DATE'] + '_' + df['TIME'] + '_' + df['SESSION GUID'] + '.html'

        book = load_workbook(ruta_archivo)
        sheet = book.active
        col_idx = sheet.max_column + 1
        sheet.cell(row=1, column=col_idx).value = 'MATCH NAME'
        for i, match_name in enumerate(df['MATCH NAME'], start=2):
            sheet.cell(row=i, column=col_idx).value = match_name

        book.save(ruta_archivo)
        messagebox.showinfo("Éxito", "Archivo procesado exitosamente y columna 'MATCH NAME' agregada.")
        return df['MATCH NAME'].tolist()

    except Exception as e:
        messagebox.showerror("Error", f"Ha ocurrido un error: {str(e)}")
        return None

# Función para procesar los archivos ZIP y renombrar los archivos HTML
def procesar_zip(ruta_zip, match_names, output_folder):
    matches_found = 0
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            with zipfile.ZipFile(ruta_zip, 'r') as main_zip:
                main_zip.extractall(temp_dir)

                for root, dirs, files in os.walk(temp_dir):
                    for file_name in files:
                        if file_name.endswith('.zip'):
                            zip_file_path_internal = os.path.join(root, file_name)
                            with zipfile.ZipFile(zip_file_path_internal, 'r') as inner_zip:
                                html_files = [f for f in inner_zip.namelist() if f.endswith('.html')]

                                for html_file_name in html_files:
                                    new_html_name = f"{os.path.splitext(file_name)[0]}.html"
                                    html_output_path = os.path.join(output_folder, new_html_name)

                                    with inner_zip.open(html_file_name) as html_file:
                                        with open(html_output_path, 'wb') as f_out:
                                            f_out.write(html_file.read())

                                    # Verificar si el nombre del archivo HTML coincide con la columna MATCH NAME
                                    if new_html_name in match_names:
                                        matches_found += 1

        if matches_found > 0:
            messagebox.showinfo("Coincidencias", f"Se encontraron {matches_found} coincidencias.")
        else:
            messagebox.showwarning("Sin coincidencias", "No se encontraron coincidencias.")
    
    except Exception as e:
        messagebox.showerror("Error", f"Hubo un problema procesando los archivos ZIP: {str(e)}")

# Funciones de selección de archivos y carpetas
def select_excel_file():
    archivo_excel = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if archivo_excel:
        excel_file_var.set(archivo_excel)
        match_names = procesar_excel(archivo_excel)
        if match_names:
            select_zip_file(match_names)

def select_zip_file(match_names):
    archivo_zip = filedialog.askopenfilename(filetypes=[("ZIP files", "*.zip")])
    if archivo_zip:
        zip_file_var.set(archivo_zip)
        output_folder = filedialog.askdirectory(title="Selecciona la carpeta de salida")
        if output_folder:
            procesar_zip(archivo_zip, match_names, output_folder)

def select_output_folder():
    folder = filedialog.askdirectory(title="Selecciona la carpeta de salida")
    if folder:
        output_folder_var.set(folder)

# Configuración de la interfaz gráfica
root = Tk()
root.title("Procesador de Archivos de Llamadas")
root.geometry("500x500")

style = Style()
style.theme_use('clam')

main_frame = Frame(root, padx=20, pady=20)
main_frame.pack(fill='both', expand=True)

excel_file_var = StringVar()
zip_file_var = StringVar()
output_folder_var = StringVar()

# Interfaz para seleccionar el Excel
Label(main_frame, text="Archivo Excel:").pack(anchor='w')
Label(main_frame, textvariable=excel_file_var, width=50, relief="sunken", padx=5).pack(fill='x', pady=(0, 5))
Button(main_frame, text="Seleccionar Excel", command=select_excel_file).pack(pady=(0, 10))

# Interfaz para seleccionar el ZIP
Label(main_frame, text="Archivo ZIP:").pack(anchor='w')
Label(main_frame, textvariable=zip_file_var, width=50, relief="sunken", padx=5).pack(fill='x', pady=(0, 5))
Button(main_frame, text="Seleccionar ZIP", command=lambda: select_zip_file(match_case)).pack(pady=(0, 10))

# Interfaz para seleccionar la carpeta de salida (sin funcionalidad)
Label(main_frame, text="Carpeta de Salida:").pack(anchor='w')
Label(main_frame, textvariable=output_folder_var, width=50, relief="sunken", padx=5).pack(fill='x', pady=(0, 5))
Button(main_frame, text="Seleccionar Carpeta", command=select_output_folder).pack(pady=(0, 10))

# Botón para procesar archivos (sin funcionalidad de momento)
Button(main_frame, text="Procesar Archivos", command=lambda: None, bg='#4CAF50', fg='white', padx=10, pady=5).pack(pady=20)

root.mainloop()
