import os
import pandas as pd
import zipfile
from tkinter import Tk, filedialog, Label, Button, Entry, messagebox, Frame, StringVar
from tkinter.ttk import Style
from datetime import datetime
import tempfile  # Para manejar directorios temporales
from openpyxl import load_workbook

# Función para formatear fechas
def parse_date(date_str):
    date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y%m%d']
    for fmt in date_formats:
        try:
            return datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None

# Función para procesar el archivo Excel
def procesar_excel(ruta_archivo):
    try:
        # Cargar el archivo Excel
        df = pd.read_excel(ruta_archivo)

        # Validar si las columnas necesarias están presentes
        columnas_necesarias = ['DATE', 'TIME', 'SESSION GUID']
        if not all(col in df.columns for col in columnas_necesarias):
            messagebox.showerror("Error", "El archivo no contiene las columnas necesarias")
            return

        # Procesar las columnas DATE y TIME
        df['DATE'] = df['DATE'].astype(str).str.replace("/", "-")
        df['TIME'] = df['TIME'].astype(str).str.replace(":", "")
        
        # Crear la columna MATCH NAME
        df['MATCH NAME'] = 'chat_' + df['DATE'] + '' + df['TIME'] + '' + df['SESSION GUID']

        # Cargar el archivo Excel con openpyxl para mantener el formato
        book = load_workbook(ruta_archivo)
        sheet = book.active

        # Escribir la nueva columna MATCH NAME en el archivo, manteniendo el formato
        col_idx = sheet.max_column + 1  # Ubicar la nueva columna al final
        sheet.cell(row=1, column=col_idx).value = 'MATCH NAME'  # Agregar el encabezado

        # Insertar los valores de la nueva columna
        for i, match_name in enumerate(df['MATCH NAME'], start=2):
            sheet.cell(row=i, column=col_idx).value = match_name

        # Guardar el archivo sobrescribiendo el original
        book.save(ruta_archivo)
        messagebox.showinfo("Éxito", "Archivo procesado exitosamente y columna 'MATCH NAME' agregada.")
    
    except Exception as e:
        messagebox.showerror("Error", f"Ha ocurrido un error: {str(e)}")

# Funciones de selección de archivos
def select_excel_file():
    global excel_file_path
    excel_file_path = filedialog.askopenfilename(
        title="Selecciona el archivo Excel",
        filetypes=[("Archivos Excel", "*.xlsx *.xls")]
    )
    if excel_file_path:
        excel_file_var.set(os.path.basename(excel_file_path))
        # Procesar el archivo Excel inmediatamente después de seleccionar
        procesar_excel(excel_file_path)

def select_zip_file():
    global zip_file_path
    zip_file_path = filedialog.askopenfilename(
        title="Selecciona el archivo ZIP",
        filetypes=[("Archivos ZIP", "*.zip")]
    )
    if zip_file_path:
        zip_file_var.set(os.path.basename(zip_file_path))

def select_output_folder():
    global output_folder_path
    output_folder_path = filedialog.askdirectory(title="Selecciona la carpeta de salida")
    if output_folder_path:
        output_folder_var.set(os.path.basename(output_folder_path))

# Función principal de procesamiento
def process_files():
    if not excel_file_path or not zip_file_path or not output_folder_path:
        messagebox.showerror("Error", "Debes seleccionar el archivo Excel, el archivo ZIP y la carpeta de salida")
        return
    
    try:
        if not os.path.exists(output_folder_path) or not os.path.isdir(output_folder_path):
            messagebox.showerror("Error", "La ruta de salida no es válida")
            return

        # Parsear la fecha si es proporcionada
        date_str = date_entry.get()
        if date_str:
            folder_date = parse_date(date_str)
            if not folder_date:
                messagebox.showerror("Error", "El formato de la fecha no es válido. Usa formatos como 'YYYY-MM-DD', 'DD/MM/YYYY', 'DD-MM-YYYY', o 'YYYYMMDD'")
                return
        else:
            folder_date = datetime.now().strftime('%Y-%m-%d')

        output_folder = os.path.join(output_folder_path, folder_date)
        os.makedirs(output_folder, exist_ok=True)

        # Crear un directorio temporal para extraer los archivos ZIP
        with tempfile.TemporaryDirectory() as temp_dir:
            # Descomprimir el archivo ZIP principal en el directorio temporal
            with zipfile.ZipFile(zip_file_path, 'r') as main_zip:
                main_zip.extractall(temp_dir)

                # Recorrer los archivos ZIP internos
                for root, dirs, files in os.walk(temp_dir):
                    for file_name in files:
                        if file_name.endswith('.zip'):
                            zip_file_path_internal = os.path.join(root, file_name)
                            with zipfile.ZipFile(zip_file_path_internal, 'r') as inner_zip:
                                html_files = [f for f in inner_zip.namelist() if f.endswith('.html')]

                                for html_file_name in html_files:
                                    # Renombrar el archivo HTML usando el nombre del archivo ZIP interno
                                    new_html_name = f"{os.path.splitext(file_name)[0]}.html"
                                    html_output_path = os.path.join(output_folder, new_html_name)

                                    with inner_zip.open(html_file_name) as html_file:
                                        with open(html_output_path, 'wb') as f_out:
                                            f_out.write(html_file.read())

        messagebox.showinfo("Éxito", "Archivos ZIP procesados y descomprimidos correctamente.")

    except Exception as e:
        messagebox.showerror("Error", f"Hubo un problema procesando los archivos: {str(e)}")

# Configuración de la interfaz gráfica
root = Tk()
root.title("Procesador de Archivos de Llamadas")
root.geometry("500x500")

style = Style()
style.theme_use('clam')

main_frame = Frame(root, padx=20, pady=20)
main_frame.pack(fill='both', expand=True)

excel_file_var = StringVar()
zip_file_var = StringVar()
output_folder_var = StringVar()

Label(main_frame, text="Archivo Excel:").pack(anchor='w')
Label(main_frame, textvariable=excel_file_var, width=50, relief="sunken", padx=5).pack(fill='x', pady=(0, 5))
Button(main_frame, text="Seleccionar Excel", command=select_excel_file).pack(pady=(0, 10))

Label(main_frame, text="Archivo ZIP:").pack(anchor='w')
Label(main_frame, textvariable=zip_file_var, width=50, relief="sunken", padx=5).pack(fill='x', pady=(0, 5))
Button(main_frame, text="Seleccionar ZIP", command=select_zip_file).pack(pady=(0, 10))

Label(main_frame, text="Carpeta de Salida:").pack(anchor='w')
Label(main_frame, textvariable=output_folder_var, width=50, relief="sunken", padx=5).pack(fill='x', pady=(0, 5))
Button(main_frame, text="Seleccionar Carpeta", command=select_output_folder).pack(pady=(0, 10))

Label(main_frame, text="Fecha (opcional, formato: YYYY-MM-DD):").pack(anchor='w')
date_entry = Entry(main_frame)
date_entry.pack(fill='x', pady=(0, 10))

Button(main_frame, text="Procesar Archivos", command=process_files, bg='#4CAF50', fg='white', padx=10, pady=5).pack(pady=20)

root.mainloop()
