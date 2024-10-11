import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

# Función para procesar el archivo Excel
def procesar_excel(ruta_archivo):
    try:
        # Cargar el archivo Excel
        df = pd.read_excel(ruta_archivo)

        # Validar si las columnas necesarias están presentes
        columnas_necesarias = ['DATE', 'TIME', 'SESSION GUID']
        if not all(col in df.columns for col in columnas_necesarias):
            messagebox.showerror("Error", "El archivo no contiene las columnas necesarias")
            return

        # Procesar las columnas DATE y TIME
        df['DATE'] = df['DATE'].astype(str).str.replace("/", "-")
        df['TIME'] = df['TIME'].astype(str).str.replace(":", "")
        
        # Crear la columna MATCH NAME
        df['MATCH NAME'] = 'chat_' + df['DATE'] + '_' + df['TIME'] + '_' + df['SESSION GUID']

        # Cargar el archivo Excel con openpyxl para mantener el formato
        book = load_workbook(ruta_archivo)
        sheet = book.active

        # Escribir la nueva columna MATCH NAME en el archivo, manteniendo el formato
        col_idx = sheet.max_column + 1  # Ubicar la nueva columna al final
        sheet.cell(row=1, column=col_idx).value = 'MATCH NAME'  # Agregar el encabezado

        # Insertar los valores de la nueva columna
        for i, match_name in enumerate(df['MATCH NAME'], start=2):
            sheet.cell(row=i, column=col_idx).value = match_name

        # Guardar el archivo sobrescribiendo el original
        book.save(ruta_archivo)
        messagebox.showinfo("Éxito", "Archivo procesado exitosamente y columna 'MATCH NAME' agregada.")
    
    except Exception as e:
        messagebox.showerror("Error", f"Ha ocurrido un error: {str(e)}")

# Función para seleccionar el archivo Excel
def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if archivo:
        procesar_excel(archivo)

# Configuración de la interfaz gráfica
root = tk.Tk()
root.title("Procesador de Excel")

label = tk.Label(root, text="Selecciona un archivo Excel para procesar")
label.pack(pady=10)

boton_seleccionar = tk.Button(root, text="Seleccionar archivo", command=seleccionar_archivo)
boton_seleccionar.pack(pady=10)

root.mainloop()
